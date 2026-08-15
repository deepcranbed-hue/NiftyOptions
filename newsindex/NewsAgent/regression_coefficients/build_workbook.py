"""Build the consolidated coefficients workbook from /tmp/coeffs.json."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C = json.load(open("/tmp/coeffs.json"))

wb = Workbook()
ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
CAL_FILL = PatternFill("solid", fgColor="E2EFDA")   # green-ish = calibrated (data)
PRIOR_FILL = PatternFill("solid", fgColor="FFF2CC")  # yellow = prior (edit these)
TITLE_FONT = Font(name=ARIAL, bold=True, size=14)
NOTE_FONT = Font(name=ARIAL, italic=True, size=9, color="595959")
BOLD = Font(name=ARIAL, bold=True, size=11)
REG = Font(name=ARIAL, size=10)
BLUE = Font(name=ARIAL, size=10, color="0000FF")     # editable input value
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr(ws, row, headers, start=1):
    for j, h in enumerate(headers, start):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = BORDER
    ws.row_dimensions[row].height = 30


def widths(ws, w):
    for i, x in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = x


def title(ws, t, sub=""):
    ws["A1"] = t; ws["A1"].font = TITLE_FONT
    if sub:
        ws["A2"] = sub; ws["A2"].font = NOTE_FONT


# ------------------------------------------------------------------ INDEX
ws = wb.active; ws.title = "Index"
title(ws, "NewsAgent — Regression & Model Coefficients",
      "Every weight/coefficient in the engine + overlay, tagged by evidence basis. "
      "Review the PRIOR (yellow) sheets; the CALIBRATED (green) sheets are data-derived.")
r = 4
ws.cell(r, 1, "Sheet").font = HDR_FONT; ws.cell(r,1).fill=HDR_FILL; ws.cell(r,1).border=BORDER
ws.cell(r, 2, "What it holds").font = HDR_FONT; ws.cell(r,2).fill=HDR_FILL; ws.cell(r,2).border=BORDER
ws.cell(r, 3, "Basis").font = HDR_FONT; ws.cell(r,3).fill=HDR_FILL; ws.cell(r,3).border=BORDER
ws.cell(r, 4, "Source").font = HDR_FONT; ws.cell(r,4).fill=HDR_FILL; ws.cell(r,4).border=BORDER
ws.row_dimensions[r].height = 24
index_rows = [
 ("Engine SENSITIVITY", "Index move per driver (Nifty, Bank Nifty)", "CALIBRATED", "calibrate.py OLS regression, ~3y (fii/geo hand-set)"),
 ("Relationship hit-rates", "Historical reliability per cross-asset relationship", "CALIBRATED", "build_events.py → events.db linkage_conf"),
 ("Event analogues", "Median outcome after a condition (e.g. sox_drop_3)", "CALIBRATED", "build_events.py → events.db event_stats"),
 ("Sector factor library", "12 sectors × their own factors, weights & signs", "PRIOR", "overlay/sector_factors.py (judgment)"),
 ("Interaction terms", "Cross-driver second-order effects", "PRIOR", "overlay/interactions.py (judgment)"),
 ("Level amplifiers", "Oil/USDINR/VIX level bands × multipliers", "PRIOR", "overlay/amplifiers.py (judgment)"),
 ("Impact baselines", "Per-metric baseline+std for surprise scoring", "PRIOR", "overlay/impact_scoring.py (judgment)"),
 ("Normalization caps", "Scale of a 'large' move per driver", "PRIOR", "overlay/common.py CAPS"),
 ("Relationship tiers", "Primary systematic drivers (market beta)", "PRIOR", "overlay/relationship_tiers.py (rule-based)"),
]
for i, (s, w, b, src) in enumerate(index_rows):
    rr = r + 1 + i
    ws.cell(rr, 1, s).font = BOLD
    ws.cell(rr, 2, w).font = REG
    cb = ws.cell(rr, 3, b); cb.font = REG
    cb.fill = CAL_FILL if b == "CALIBRATED" else PRIOR_FILL
    ws.cell(rr, 4, src).font = NOTE_FONT
    for cc in range(1,5): ws.cell(rr,cc).border=BORDER; ws.cell(rr,cc).alignment=Alignment(wrap_text=True, vertical="center")
ll = r + len(index_rows) + 2
ws.cell(ll, 1, "Legend:").font = BOLD
ws.cell(ll+1, 1, "  CALIBRATED (green) = fitted from history — trust it. Refresh via calibrate.py / build_events.py.").font = NOTE_FONT
ws.cell(ll+2, 1, "  PRIOR (yellow) = judgment weight — reasonable but not fitted; edit here to tune, then reload the overlay.").font = NOTE_FONT
ws.cell(ll+3, 1, "  Blue values = the numbers you can change.").font = NOTE_FONT
widths(ws, [26, 46, 14, 52])

# ------------------------------------------------- ENGINE SENSITIVITY
ws = wb.create_sheet("Engine SENSITIVITY")
title(ws, "Engine SENSITIVITY (index % move per unit driver)",
      "CALIBRATED — OLS regression over ~3y (calibrate.py). fii_kcr & geopolitics_hits are hand-set (no historical series).")
sens = C["sensitivity"]; indices = list(sens.keys())
drivers = list(sens[indices[0]].keys())
hdr(ws, 4, ["Driver"] + indices + ["Basis"])
for i, d in enumerate(drivers):
    rr = 5 + i
    ws.cell(rr, 1, d).font = BOLD
    for j, idx in enumerate(indices, 2):
        c = ws.cell(rr, j, round(sens[idx][d], 4)); c.font = BLUE; c.border = BORDER; c.number_format = "0.0000"
    handset = d in ("fii_kcr", "geopolitics_hits")
    cb = ws.cell(rr, 2 + len(indices), "hand-set" if handset else "fitted (regression)")
    cb.font = REG; cb.fill = PRIOR_FILL if handset else CAL_FILL; cb.border = BORDER
    ws.cell(rr,1).border=BORDER
widths(ws, [22] + [14]*len(indices) + [22])

# ------------------------------------------------- RELATIONSHIP HIT-RATES
ws = wb.create_sheet("Relationship hit-rates")
title(ws, "Relationship hit-rates (historical reliability)",
      "CALIBRATED — from events.db (build_events.py). hit-rate = fraction of days the proxy moved as the relationship predicts.")
hdr(ws, 4, ["Relationship", "Hit-rate %", "Sample n", "Tag"])
for i, (name, v) in enumerate(sorted(C["linkage_conf"].items(), key=lambda kv: -kv[1]["hit_rate"])):
    rr = 5 + i
    ws.cell(rr, 1, name).font = REG
    ws.cell(rr, 2, v["hit_rate"]).font = BLUE; ws.cell(rr,2).number_format = "0"
    ws.cell(rr, 3, v["n"]).font = REG
    tag = "CALIBRATED" if v["n"] >= 60 else "PRIOR"
    ct = ws.cell(rr, 4, tag); ct.fill = CAL_FILL if tag=="CALIBRATED" else PRIOR_FILL
    for cc in range(1,5): ws.cell(rr,cc).border=BORDER
widths(ws, [46, 12, 12, 14])

# ------------------------------------------------- EVENT ANALOGUES
ws = wb.create_sheet("Event analogues")
title(ws, "Event analogues (historical outcome after a condition)",
      "CALIBRATED — events.db event_stats. Median % move of each target in the sessions following the condition.")
es = C.get("event_stats", {})
targets = sorted({t for v in es.values() for t in v})
hdr(ws, 4, ["Condition"] + [f"{t} median%" for t in targets] + ["n"])
for i, (cond, tv) in enumerate(es.items()):
    rr = 5 + i
    ws.cell(rr, 1, cond).font = BOLD
    n = 0
    for j, t in enumerate(targets, 2):
        val = tv.get(t, {})
        ws.cell(rr, j, val.get("median")).font = REG; ws.cell(rr,j).number_format="0.00"
        n = val.get("n", n)
    ws.cell(rr, 2+len(targets), n).font = REG
    for cc in range(1, 3+len(targets)): ws.cell(rr,cc).border=BORDER
widths(ws, [18] + [16]*len(targets) + [8])

# ------------------------------------------------- SECTOR FACTOR LIBRARY
ws = wb.create_sheet("Sector factor library")
title(ws, "Per-sector factor library (each sector its OWN drivers)",
      "PRIOR — judgment weights (overlay/sector_factors.py). sign = effect on the sector when the signal is positive. Edit the yellow weights.")
hdr(ws, 4, ["Sector", "Factor", "Weight", "Sign", "Kind", "Signal key"])
rr = 5
for sector, rows in C["sector_library"].items():
    for k, f in enumerate(rows):
        ws.cell(rr, 1, sector if k == 0 else "").font = BOLD
        ws.cell(rr, 2, f["factor"]).font = REG
        cw = ws.cell(rr, 3, f["weight"]); cw.font = BLUE; cw.fill = PRIOR_FILL; cw.number_format = "0.00"
        ws.cell(rr, 4, "＋" if f["sign"] > 0 else "－").font = REG
        ws.cell(rr, 5, f["kind"]).font = REG
        ws.cell(rr, 6, str(f["source"])).font = NOTE_FONT
        for cc in range(1,7): ws.cell(rr,cc).border=BORDER
        rr += 1
widths(ws, [22, 40, 10, 8, 14, 22])

# ------------------------------------------------- INTERACTIONS
ws = wb.create_sheet("Interaction terms")
title(ws, "Cross-driver interaction terms (second-order effects)",
      "PRIOR — overlay/interactions.py. magnitude = |normA| × |normB| × weight; sign set by the rule.")
hdr(ws, 4, ["Term", "Leg A", "Leg B", "Weight", "Rule"])
for i, t in enumerate(C["interactions"]):
    rr = 5 + i
    ws.cell(rr, 1, t["term"]).font = BOLD
    ws.cell(rr, 2, t["leg_a"]).font = REG
    ws.cell(rr, 3, t["leg_b"]).font = REG
    cw = ws.cell(rr, 4, t["weight"]); cw.font = BLUE; cw.fill = PRIOR_FILL; cw.number_format = "0.00"
    ws.cell(rr, 5, t["rule"]).font = REG
    for cc in range(1,6): ws.cell(rr,cc).border=BORDER
widths(ws, [22, 16, 18, 10, 22])

# ------------------------------------------------- LEVEL AMPLIFIERS
ws = wb.create_sheet("Level amplifiers")
title(ws, "Level amplifiers (same % move bites differently by level)",
      "PRIOR — overlay/amplifiers.py. multiplier scales a driver's impact by its absolute level band.")
rr = 4
for label, key in [("Oil (Brent $)", "oil_bands"), ("USDINR (₹)", "usdinr_bands"), ("India VIX", "vix_bands")]:
    ws.cell(rr, 1, label).font = BOLD; rr += 1
    hdr(ws, rr, ["Band", "Multiplier", "From level ≥"]); rr += 1
    for b in C[key]:
        ws.cell(rr, 1, b["band"]).font = REG
        cm = ws.cell(rr, 2, b["multiplier"]); cm.font = BLUE; cm.fill = PRIOR_FILL; cm.number_format = "0.0"
        ws.cell(rr, 3, b["from_level"]).font = REG
        for cc in range(1,4): ws.cell(rr,cc).border=BORDER
        rr += 1
    rr += 1
widths(ws, [30, 12, 14])

# ------------------------------------------------- IMPACT BASELINES
ws = wb.create_sheet("Impact baselines")
title(ws, "Impact-scoring baselines (surprise = σ from baseline)",
      "PRIOR — overlay/impact_scoring.py. A metric's surprise = (value − baseline) / std.")
hdr(ws, 4, ["Metric", "Baseline", "Std", "Higher is better?", "Unit"])
for i, (m, v) in enumerate(C["impact_baselines"].items()):
    rr = 5 + i
    ws.cell(rr, 1, m).font = BOLD
    cb = ws.cell(rr, 2, v["baseline"]); cb.font = BLUE; cb.fill = PRIOR_FILL; cb.number_format = "0.0"
    cs = ws.cell(rr, 3, v["std"]); cs.font = BLUE; cs.fill = PRIOR_FILL; cs.number_format = "0.0"
    ws.cell(rr, 4, "yes" if v["higher_is_better"] else "no").font = REG
    ws.cell(rr, 5, v["unit"]).font = REG
    for cc in range(1,6): ws.cell(rr,cc).border=BORDER
widths(ws, [26, 12, 10, 18, 8])

# ------------------------------------------------- CAPS
ws = wb.create_sheet("Normalization caps")
title(ws, "Normalization caps (scale of a 'large' move)",
      "overlay/common.py CAPS — used to map a raw move into [-1, 1] before weighting.")
hdr(ws, 4, ["Driver key", "Cap"])
for i, (k, v) in enumerate(C["caps"].items()):
    rr = 5 + i
    ws.cell(rr, 1, k).font = BOLD
    cc = ws.cell(rr, 2, v); cc.font = BLUE; cc.fill = PRIOR_FILL; cc.number_format = "0.0"
    ws.cell(rr,1).border=BORDER; ws.cell(rr,2).border=BORDER
widths(ws, [22, 10])

# ------------------------------------------------- TIERS PRIMARY
ws = wb.create_sheet("Relationship tiers")
title(ws, "Relationship tiers — PRIMARY (systematic / market beta)",
      "PRIOR classification (overlay/relationship_tiers.py). sign = +1 means a positive value pushes the whole market UP.")
hdr(ws, 4, ["Relationship", "Driver key", "Sign", "Mechanism"])
for i, t in enumerate(C["tiers_primary"]):
    rr = 5 + i
    ws.cell(rr, 1, t["relationship"]).font = BOLD
    ws.cell(rr, 2, t["driver"]).font = REG
    ws.cell(rr, 3, "＋" if t["sign"] > 0 else "－").font = REG
    ws.cell(rr, 4, t["mechanism"]).font = REG
    for cc in range(1,5): ws.cell(rr,cc).border=BORDER; ws.cell(rr,cc).alignment=Alignment(wrap_text=True, vertical="center")
widths(ws, [26, 18, 8, 60])

for s in wb.sheetnames:
    wb[s].sheet_view.showGridLines = False
    wb[s].freeze_panes = "A5" if s != "Index" else "A5"

out = "/sessions/gifted-jolly-archimedes/mnt/newsindex/NewsAgent/regression_coefficients/NewsAgent_coefficients.xlsx"
wb.save(out)
print("saved", out, "| sheets:", wb.sheetnames)
