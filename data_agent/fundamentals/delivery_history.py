#!/usr/bin/env python3
"""delivery_history.py — has this company actually DELIVERED, year after year?

WHY THIS EXISTS RATHER THAN REUSING eps_history.json
    eps_history.json reads stored EPS out of Postgres by matching NINE label
    variants ("eps", "basic eps", "diluted eps", "earnings per share", ...). Nothing
    forces one variant to win consistently, so different years in one series can come
    from different definitions. The result is not noisy, it is wrong by multiples:

        HDFC Bank FY23   stored 82.6   net profit / shares = 39.5   (2.1x)
        Kotak     FY23   stored 75.0   net profit / shares = 11.0   (6.8x)
        Bajaj Fin FY24   stored 236.0  net profit / shares = 23.3   (10.1x)

    HDFC Bank's first six and last two years reconcile exactly; two years in the
    middle do not. A screen built on that throws HDFC Bank out for a "-48.7% year"
    during which net profit grew 11 percent.

    So this file does not read EPS at all. It reads two unambiguous single-label
    series and divides them.

WHY NET PROFIT IS THE CONSISTENCY MEASURE, NOT EPS
    EPS moves when the share count moves. HDFC Bank issued ~400 crore shares to
    absorb HDFC Ltd; Bajaj Finance has bonused and split. Judging "did the business
    deliver" on EPS penalises a company for issuing paper to buy something, and
    rewards one for buying its own stock back. Net profit answers the delivery
    question; derived EPS answers the per-share question. Both are reported here,
    and any year where the share count moved more than SHARE_MOVE_FLAG is marked so
    a merger shows up as a merger instead of a collapse.

SOURCE
    The Screener exports in screener_data/, read from the 'Data Sheet' tab — the raw
    values. The visible 'Profit & Loss' tab is a formula view whose cached values are
    empty unless Excel last saved the file, which is why a naive reader concludes the
    export is blank.

    This is the same upstream that feeds Postgres fundamentals.financials via
    ingest_screener.py. It is read directly here for one reason: the label ambiguity
    above is introduced on the Postgres side, and the raw sheet has exactly one row
    called 'Net profit' and one called 'Adjusted Equity Shares in Cr'. No second
    table is created — the output is a derived JSON artifact, the same convention as
    pe_history.json, eps_history.json and fii_holdings.json.
"""
from __future__ import annotations

import csv
import datetime
import json
import os
import sys
import warnings

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(os.path.dirname(_HERE))
DATA = os.path.join(_HERE, "screener_data")
OUT = os.path.join(_ROOT, "delivery_history.json")

# A year-on-year change is only meaningful if the base is a real profit. Below this
# (Rs crore) the ratio explodes and reports things like -7541%.
_PROFIT_FLOOR = 1.0
# Share count moving more than this in one year is a corporate action, not organic.
SHARE_MOVE_FLAG = 5.0
# A share count this far below the series median is not a real count. Zomato's early
# rows carry 0.03 cr against ~900 cr actual, which turns EPS into -78,905. Derived EPS
# is only as good as its denominator, so an implausible denominator must yield None
# rather than a confident absurdity.
_SHARE_SANITY_RATIO = 0.10
# A growth rate off a near-zero base is arithmetically true and analytically useless:
# Axis Bank earned Rs 276 cr in FY18 after provisions and Rs 4,677 cr in FY19, which is
# +1596 percent and means "recovered", not "grew sixteenfold". Flag it so the reader
# sees a recovery rather than a trend.
_DEPRESSED_BASE_RATIO = 0.25
# Fewer periods than this and "consistency" is an anecdote.
_MIN_YEARS = 6


def _is_date(v) -> bool:
    return isinstance(v, (datetime.datetime, datetime.date))


def _grid(sym: str):
    p = os.path.join(DATA, f"{sym}.xlsx")
    if not os.path.exists(p):
        return None
    wb = openpyxl.load_workbook(p, data_only=False, read_only=True)
    if "Data Sheet" not in wb.sheetnames:
        wb.close()
        return None
    rows = [list(r) for r in wb["Data Sheet"].iter_rows(max_col=20, values_only=True)]
    wb.close()
    return rows


def _find(rows, text, start=0):
    t = text.strip().upper()
    for i in range(start, len(rows)):
        r = rows[i]
        if r and r[0] and str(r[0]).strip().upper().startswith(t):
            return i
    return None


def _row_after(rows, label, lo, hi):
    """First row named `label` strictly inside [lo, hi). Bounds matter: 'Net profit'
    and 'Sales' each appear in BOTH the annual block and the quarterly block, and
    reading the wrong one silently mixes frequencies — the same defect that put a
    June quarter into every IT company's annual series."""
    for i in range(lo, min(hi, len(rows))):
        r = rows[i]
        if r and r[0] and str(r[0]).strip().lower() == label.lower():
            return [c if isinstance(c, (int, float)) else None for c in r[1:]]
    return None


def _dates(rows, idx):
    return [c if _is_date(c) else None for c in rows[idx][1:]]


def _cagr(a, b, years):
    if a is None or b is None or a <= 0 or b <= 0 or years <= 0:
        return None
    return round(((b / a) ** (1.0 / years) - 1.0) * 100.0, 1)


def _quarters(rows, q, end) -> list[dict]:
    """The QUARTERS block, as-is. Ten quarters is what Screener exports.

    Kept deliberately separate from the annual series rather than merged into it. The
    two answer different questions — the annual block is whether the business delivered,
    the quarterly block is whether it is delivering NOW — and mixing frequencies in one
    series is the exact defect that put a June quarter into every IT major's annual EPS
    history. YoY is quarter-over-same-quarter-last-year (lag 4), never sequential:
    Indian earnings are seasonal enough that Q1-vs-Q4 is mostly the calendar.
    """
    if q is None:
        return []
    dhdr = _find(rows, "Report Date", q)
    if dhdr is None or dhdr >= end:
        return []
    dates = _dates(rows, dhdr)
    sales = _row_after(rows, "Sales", q, end)
    profit = _row_after(rows, "Net profit", q, end)
    op = _row_after(rows, "Operating Profit", q, end)
    if not profit:
        return []
    qs = []
    for i, d in enumerate(dates):
        if not _is_date(d):
            continue
        np_ = profit[i] if i < len(profit) else None
        if np_ is None:
            continue
        qs.append({
            "period": d.strftime("%Y-%m-%d"),
            "sales": round(sales[i], 1) if (sales and i < len(sales) and sales[i] is not None) else None,
            "net_profit": round(np_, 1),
            "operating_profit": round(op[i], 1) if (op and i < len(op) and op[i] is not None) else None,
            "yoy_pct": None,
        })
    for i in range(4, len(qs)):
        base = qs[i - 4]["net_profit"]
        if base is not None and base >= _PROFIT_FLOOR:
            qs[i]["yoy_pct"] = round((qs[i]["net_profit"] / base - 1.0) * 100.0, 1)
    return qs


def build(sym: str) -> dict | None:
    rows = _grid(sym)
    if not rows:
        return None
    pl = _find(rows, "PROFIT & LOSS")
    q = _find(rows, "QUARTERS")
    bs = _find(rows, "BALANCE SHEET")
    der = _find(rows, "DERIVED")
    if pl is None or bs is None:
        return None
    ann_end = q if (q is not None and q > pl) else bs

    dhdr = _find(rows, "Report Date", pl)
    if dhdr is None:
        return None
    dates = _dates(rows, dhdr)
    sales = _row_after(rows, "Sales", pl, ann_end)
    profit = _row_after(rows, "Net profit", pl, ann_end)
    if not profit:
        return None

    # Share count: 'Adjusted Equity Shares in Cr' lives in DERIVED and is already
    # normalised for splits/bonuses. 'No. of Equity Shares' in the balance sheet is not.
    shares = None
    if der is not None:
        shares = _row_after(rows, "Adjusted Equity Shares in Cr", der, len(rows))
    if not shares:
        shares = _row_after(rows, "No. of Equity Shares", bs, der or len(rows))

    # Median share count first: it is the reference for spotting a bad denominator.
    _sh_vals = [x for x in (shares or []) if isinstance(x, (int, float)) and x > 0]
    _sh_med = sorted(_sh_vals)[len(_sh_vals) // 2] if _sh_vals else None

    series = []
    for i, d in enumerate(dates):
        if not _is_date(d):
            continue
        np_ = profit[i] if i < len(profit) else None
        if np_ is None:
            continue
        sh = shares[i] if (shares and i < len(shares)) else None
        if sh and _sh_med and sh < _sh_med * _SHARE_SANITY_RATIO:
            sh = None       # implausible denominator -> no EPS, not a fake one
        series.append({
            "period": d.strftime("%Y-%m-%d"),
            "sales": round(sales[i], 1) if (sales and i < len(sales) and sales[i] is not None) else None,
            "net_profit": round(np_, 1),
            "shares_cr": round(sh, 2) if sh else None,
            # The whole point of this file: EPS derived, never read.
            "eps": round(np_ / sh, 2) if (sh and sh > 0) else None,
        })
    if len(series) < _MIN_YEARS:
        return None

    # ---- delivery consistency, measured on the PATH not the endpoints ----------
    yoy, flags = [], []
    for i in range(1, len(series)):
        a, b = series[i - 1]["net_profit"], series[i]["net_profit"]
        if a is not None and b is not None and a >= _PROFIT_FLOOR:
            yoy.append(round((b / a - 1.0) * 100.0, 1))
        else:
            yoy.append(None)          # loss base: undefined, never a huge number
        s0, s1 = series[i - 1]["shares_cr"], series[i]["shares_cr"]
        moved = (abs(s1 / s0 - 1.0) * 100.0) if (s0 and s1) else 0.0
        flags.append(round(moved, 1) if moved > SHARE_MOVE_FLAG else None)

    # Which YoY figures sit on a depressed base? Compare each base against the median
    # profit of the series — a base far below it makes the percentage a recovery
    # statistic, not a growth rate.
    _p = [s_["net_profit"] for s_ in series if s_["net_profit"] and s_["net_profit"] > 0]
    _p_med = sorted(_p)[len(_p) // 2] if _p else None
    depressed = []
    for i in range(1, len(series)):
        base = series[i - 1]["net_profit"]
        depressed.append(bool(_p_med and base is not None
                              and 0 < base < _p_med * _DEPRESSED_BASE_RATIO))

    real = [x for x in yoy if x is not None]
    grew = sum(1 for x in real if x > 0)
    span = (datetime.date.fromisoformat(series[-1]["period"])
            - datetime.date.fromisoformat(series[0]["period"])).days / 365.25

    def window(n):
        cut = datetime.date.fromisoformat(series[-1]["period"]) - datetime.timedelta(days=int(n * 365.25))
        older = [s for s in series if datetime.date.fromisoformat(s["period"]) <= cut]
        if not older:
            return None, None
        st = older[-1]
        yrs = (datetime.date.fromisoformat(series[-1]["period"])
               - datetime.date.fromisoformat(st["period"])).days / 365.25
        return st, round(yrs, 1)

    out = {"symbol": sym, "n_years": len(series),
           "first": series[0]["period"], "last": series[-1]["period"],
           "years_grown": grew, "years_measured": len(real),
           "consistency_pct": round(grew / len(real) * 100, 1) if real else None,
           "worst_year_pct": min(real) if real else None,
           "best_year_pct": max(real) if real else None,
           "yoy_net_profit_pct": yoy,
           "depressed_base": depressed,
           "share_move_flags": flags,
           "corporate_action_years": sum(1 for f in flags if f),
           "series": series,
           "quarters": _quarters(rows, q, bs if (bs and q and bs > q) else len(rows))}
    for n in (3, 5):
        st, yrs = window(n)
        out[f"profit_cagr_{n}y_pct"] = _cagr(st["net_profit"], series[-1]["net_profit"], yrs) if st else None
        out[f"eps_cagr_{n}y_pct"] = (_cagr(st.get("eps"), series[-1].get("eps"), yrs)
                                     if st and st.get("eps") and series[-1].get("eps") else None)
    out["profit_cagr_full_pct"] = _cagr(series[0]["net_profit"], series[-1]["net_profit"], span)
    return out


def main() -> int:
    with open(os.path.join(_ROOT, "nifty-50-stock-list.csv")) as f:
        rdr = list(csv.DictReader(f))
    syms = [r["Symbol"] for r in rdr]
    sect = {r["Symbol"]: r["Sector"] for r in rdr}
    wt = {r["Symbol"]: float(r["Weight"]) for r in rdr}

    hist, missing = {}, []
    for s in syms:
        d = build(s)
        if d is None:
            missing.append(s)
            continue
        d["sector"] = sect[s]
        d["weight_pct"] = wt[s]
        hist[s] = d

    doc = {"as_of": datetime.date.today().isoformat(),
           "source": "screener_data/*.xlsx 'Data Sheet' (raw). EPS DERIVED as "
                     "net_profit / adjusted_shares — stored EPS is not used, see module docstring.",
           "note": "Consistency is measured on NET PROFIT year-on-year, which is immune to "
                   "splits, bonuses and merger dilution. share_move_flags marks any year the "
                   "share count moved more than 5 percent so a corporate action is visible "
                   "rather than scored as a collapse.",
           "history": hist}
    with open(OUT, "w") as f:
        json.dump(doc, f, indent=1)

    print(f"delivery_history.json: {len(hist)}/{len(syms)} symbols")
    if missing:
        print(f"  no usable export: {', '.join(missing)}")
    ca = sum(v["corporate_action_years"] for v in hist.values())
    print(f"  corporate-action years flagged: {ca}")
    n = [v["n_years"] for v in hist.values()]
    print(f"  periods per symbol: min {min(n)} max {max(n)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
