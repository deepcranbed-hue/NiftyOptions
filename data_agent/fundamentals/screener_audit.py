#!/usr/bin/env python3
"""screener_audit.py — what Screener data do we actually have, per Nifty 50 name?

Run this BEFORE and AFTER download_screener.py. It answers the only question that
matters before an ingest: which symbols have enough history to compute a multi-year
growth series, and which are thin or absent.

WHY IT EXISTS SEPARATELY FROM THE DOWNLOADER
    download_screener.py skips a symbol when the FILE EXISTS. That is the right
    default — the exports are rate-limited and mostly static — but it means a file
    that downloaded badly (headers present, values missing) is skipped forever. A
    file-existence check and a data-completeness check are different questions, and
    conflating them is how a half-empty export survives three re-runs.

WHAT COUNTS AS USABLE
    The visible sheets ('Profit & Loss', 'Quarters') are FORMULA VIEWS pointing at
    'Data Sheet'. Opening them with data_only=True returns None for every cell
    unless Excel itself last saved the file, so a naive reader concludes the export
    is empty when it is fine. This reads 'Data Sheet' directly, which holds the raw
    numbers, and counts real date headers and real numeric Sales points.

    Thin is not always broken: a company demerged in 2023 or one that changed its
    financial year genuinely has fewer annual periods. The report flags thin rows;
    it does not assume they are download failures.
"""
from __future__ import annotations

import csv
import datetime
import os
import sys
import warnings

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(os.path.dirname(_HERE))
DATA = os.path.join(_HERE, "screener_data")
UNIVERSE_CSV = os.path.join(_ROOT, "nifty-50-stock-list.csv")

# Below this many annual periods a CAGR is a two-point line, not a trend.
MIN_ANNUAL = 5


def _symbols() -> list[str]:
    with open(UNIVERSE_CSV) as f:
        return [r["Symbol"] for r in csv.DictReader(f)]


def _is_date(v) -> bool:
    return isinstance(v, (datetime.datetime, datetime.date))


def audit_one(sym: str) -> dict:
    path = os.path.join(DATA, f"{sym}.xlsx")
    if not os.path.exists(path):
        return {"sym": sym, "state": "MISSING"}
    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
        if "Data Sheet" not in wb.sheetnames:
            wb.close()
            return {"sym": sym, "state": "NO DATA SHEET"}
        rows = [list(r) for r in wb["Data Sheet"].iter_rows(max_col=16, values_only=True)]
        wb.close()

        def section(label):
            for i, r in enumerate(rows):
                if r and r[0] and str(r[0]).strip().upper().startswith(label):
                    return i
            return None

        pl, q = section("PROFIT & LOSS"), section("QUARTERS")
        ann = qtr = 0
        last = None
        if pl is not None and pl + 1 < len(rows):
            ds = [c for c in rows[pl + 1][1:] if _is_date(c)]
            ann, last = len(ds), (max(ds).strftime("%Y-%m-%d") if ds else None)
        if q is not None and q + 1 < len(rows):
            qtr = len([c for c in rows[q + 1][1:] if _is_date(c)])

        # Header dates alone prove nothing — a broken export keeps its headers and
        # loses its numbers. Count actual numeric Sales points.
        sales = 0
        if pl is not None:
            for r in rows[pl: pl + 10]:
                if r and r[0] and str(r[0]).strip() == "Sales":
                    sales = len([c for c in r[1:] if isinstance(c, (int, float))])
                    break

        state = "OK" if (ann >= MIN_ANNUAL and sales >= MIN_ANNUAL) else "THIN"
        return {"sym": sym, "state": state, "annual": ann, "quarters": qtr,
                "sales_pts": sales, "latest": last}
    except Exception as e:
        return {"sym": sym, "state": f"ERROR {type(e).__name__}"}


def main() -> int:
    syms = _symbols()
    res = [audit_one(s) for s in syms]
    print(f"{'symbol':13s}{'state':>16}{'annual':>8}{'qtrs':>6}{'sales':>7}{'latest':>13}")
    for r in res:
        print(f"{r['sym']:13s}{r['state']:>16}"
              f"{r.get('annual', ''):>8}{r.get('quarters', ''):>6}"
              f"{r.get('sales_pts', ''):>7}{str(r.get('latest') or ''):>13}")

    ok = [r["sym"] for r in res if r["state"] == "OK"]
    thin = [r["sym"] for r in res if r["state"] == "THIN"]
    gone = [r["sym"] for r in res if r["state"] not in ("OK", "THIN")]
    print(f"\nOK {len(ok)}/{len(syms)}   THIN {len(thin)}   MISSING/BROKEN {len(gone)}")
    if thin:
        print(f"  THIN    : {', '.join(sorted(thin))}")
        print("            delete these to force a re-download — but check first whether the")
        print("            company simply has less history (recent demerger, financial-year change).")
    if gone:
        print(f"  MISSING : {', '.join(sorted(gone))}")
        print("            download_screener.py will fetch these; it skips everything already present.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
